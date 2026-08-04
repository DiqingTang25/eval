"""
黄金QA生成器

流程: 解析Excel外部资料 → 提取结构化文本 → LLM严格约束生成 → 输出qa_pending.json
"""

import os
import json
import uuid
from datetime import datetime
from openai import OpenAI


class ExcelExtractor:
    """从课程 Excel 提取结构化文本"""

    def __init__(self, external_dir="external_materials"):
        self.external_dir = external_dir

    def extract_all(self) -> dict[str, list[dict]]:
        """
        解析所有 Excel 文件，按 Phase 组织内容

        :return: { "PHASE 01": [{"text":..., "source_file":..., "source_sheet":...}, ...], ... }
        """
        import openpyxl

        phase_map = {
            "PHASE 01": [],
            "PHASE 02": [],
            "PHASE 03": [],
            "PHASE 04": [],
            "PHASE 05": [],
        }

        phase_keywords = {
            "PHASE 01": ["国产AI技术基础", "大模型", "HiAgent", "Agent", "Prompt", "ESP32", "云边协同", "3D建模", "Blender"],
            "PHASE 02": ["新型硬件设计", "增材制造", "减材制造", "激光雕刻", "CNC", "Arduino", "数控"],
            "PHASE 03": ["环境感知", "传感器", "摄像头", "Edge Impulse", "边缘AI", "音频", "嵌入式部署"],
            "PHASE 04": ["触觉反馈", "舵机", "灯带", "电机", "触摸", "具身", "执行器"],
            "PHASE 05": ["具身智能", "M5Stack", "StackChan", "机器人", "路演", "传感器融合", "多模态"],
        }

        for fname in os.listdir(self.external_dir):
            if not fname.endswith(".xlsx"):
                continue
            fpath = os.path.join(self.external_dir, fname)
            try:
                wb = openpyxl.load_workbook(fpath, data_only=True)
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    rows = list(ws.iter_rows(values_only=True))
                    if not rows:
                        continue

                    # 提取所有文本
                    all_text = []
                    for row in rows:
                        for cell in row:
                            if cell and str(cell).strip() and len(str(cell).strip()) > 3:
                                all_text.append(str(cell).strip())

                    combined = " ".join(all_text)
                    if len(combined) < 50:
                        continue

                    # 按关键词分到 Phase
                    assigned = False
                    for phase, keywords in phase_keywords.items():
                        if any(kw in combined for kw in keywords):
                            phase_map[phase].append({
                                "text": combined[:3000],
                                "source_file": fname,
                                "source_sheet": sheet_name,
                            })
                            assigned = True

                    # 如果没有匹配到任何 phase，存入 PHASE 01（通用）
                    if not assigned:
                        phase_map["PHASE 01"].append({
                            "text": combined[:3000],
                            "source_file": fname,
                            "source_sheet": sheet_name,
                        })

                wb.close()
            except Exception as e:
                print(f"  ⚠️ 解析 {fname} 失败: {e}")

        return phase_map


class QAGenerator:
    """黄金QA生成器"""

    TYPES = ["概念解释", "操作步骤", "对比分析", "应用场景"]
    PHASES = ["PHASE 01", "PHASE 02", "PHASE 03", "PHASE 04", "PHASE 05"]

    def __init__(self, api_key, base_url="https://api.deepseek.com/v1"):
        self.client = OpenAI(api_key=api_key, base_url=base_url)
        self.extractor = ExcelExtractor()

    def generate_from_excel(self, questions_per_phase: int = 4) -> list[dict]:
        """
        从 Excel 资料生成 QA pairs

        :param questions_per_phase: 每阶段每种类型生成的问题数
        :return: QA pair 列表
        """
        print("📚 解析外部资料...")
        phase_texts = self.extractor.extract_all()

        all_qa = []
        total = 0

        for phase in self.PHASES:
            sources = phase_texts.get(phase, [])
            if not sources:
                print(f"  ⚠️ {phase}: 无匹配内容，跳过")
                continue

            print(f"\n📝 {phase} ({len(sources)} 个文本段)")

            for q_type in self.TYPES:
                # 取该 phase 的第一个源文本（最长/最相关）
                source = sources[0]
                source_text = source["text"]

                if len(source_text) < 100:
                    continue

                print(f"   生成 {q_type}...")

                # 调用 LLM
                qa_pair = self._generate_one(
                    phase=phase,
                    q_type=q_type,
                    source_text=source_text,
                    source_file=source["source_file"],
                    source_sheet=source["source_sheet"],
                )

                if qa_pair:
                    all_qa.append(qa_pair)
                    total += 1

        print(f"\n✅ 共生成 {total} 个 QA pairs")
        return all_qa

    def _generate_one(self, phase, q_type, source_text, source_file, source_sheet) -> dict | None:
        """生成单个 QA pair"""
        from src.qa_schema import QA_GENERATION_PROMPT

        prompt = QA_GENERATION_PROMPT.format(
            phase=phase,
            type=q_type,
            difficulty="中等",
            source_text=source_text[:2500],
        )

        try:
            response = self.client.chat.completions.create(
                model="deepseek-chat",
                messages=[{"role": "user", "content": prompt}],
                temperature=0.5,
                response_format={"type": "json_object"},
            )
            data = json.loads(response.choices[0].message.content)

            # 构建标准 QA pair
            qa_id = f"QA_{phase.replace(' ', '')}_{uuid.uuid4().hex[:6].upper()}"
            return {
                "qa_id": qa_id,
                "phase": phase,
                "type": q_type,
                "difficulty": "中等",
                "question": data.get("question", ""),
                "golden_answer": data.get("golden_answer", ""),
                "knowledge_points": data.get("knowledge_points", []),
                "source": {
                    "document": source_file,
                    "sheet": source_sheet,
                    "excerpt": data.get("source_excerpt", ""),
                },
                "status": "pending",
                "reviewer_notes": "",
                "created_at": datetime.now().isoformat(),
            }

        except Exception as e:
            print(f"      ❌ 生成失败: {e}")
            return None

    def save_pending(self, qa_pairs: list[dict], output_path="data/qa_pending.json"):
        """保存待审核 QA 到文件"""
        os.makedirs(os.path.dirname(output_path), exist_ok=True)

        # 合并已有 pending（避免覆盖）
        existing = []
        if os.path.exists(output_path):
            try:
                with open(output_path, "r", encoding="utf-8") as f:
                    existing = json.load(f)
            except Exception:
                pass

        # 去重（相同 phase + question 的不重复添加）
        seen = {(qa["phase"], qa["question"]) for qa in existing}
        new_qa = [qa for qa in qa_pairs if (qa["phase"], qa["question"]) not in seen]

        all_qa = existing + new_qa
        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(all_qa, f, ensure_ascii=False, indent=2)

        print(f"💾 保存 {len(new_qa)} 条新QA（共 {len(all_qa)} 条）→ {output_path}")
        return output_path


# ── 命令行入口 ──────────────────────────────────

if __name__ == "__main__":
    import sys
    sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

    from dotenv import load_dotenv
    load_dotenv()

    api_key = os.getenv("OPENAI_API_KEY")
    if not api_key:
        print("❌ 请设置 OPENAI_API_KEY")
        sys.exit(1)

    gen = QAGenerator(api_key)
    qa_pairs = gen.generate_from_excel(questions_per_phase=4)
    gen.save_pending(qa_pairs)

    print("\n📊 生成统计:")
    for phase in QAGenerator.PHASES:
        count = sum(1 for q in qa_pairs if q["phase"] == phase)
        print(f"  {phase}: {count} 条")
